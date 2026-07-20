"""
Tests der zentralen Textverwaltung (texte.py): Laden/Zusammenfuehren
der YAML-Dateien je Sprache, Schluessel-Fallback (Zielsprache -> Deutsch
-> Schluesselname), Platzhalter-Formatierung und die Excel-spezifische
Sicht excel_texte() ohne "excel."-Prefix.
"""

from __future__ import annotations

import importlib
import os
from pathlib import Path

import pytest


@pytest.fixture(autouse=True)
def _frische_sprache():
    """Sprachwahl je Test isolieren (Umgebungsvariable + Cache)."""
    alt = os.environ.get("NOBIS_SPRACHE")
    yield
    if alt is None:
        os.environ.pop("NOBIS_SPRACHE", None)
    else:
        os.environ["NOBIS_SPRACHE"] = alt
    import texte

    texte.lade_texte.cache_clear()


def _setze_sprache(sprache: str | None):
    import texte

    if sprache is None:
        os.environ.pop("NOBIS_SPRACHE", None)
    else:
        os.environ["NOBIS_SPRACHE"] = sprache
    texte.lade_texte.cache_clear()
    importlib.reload(texte)
    return texte


class TestSpracheUndFallback:
    def test_standardsprache_deutsch(self):
        t = _setze_sprache(None)
        assert t.aktive_sprache() == "de"
        assert t.txt("oberflaeche.nav_portfolio") == "Portfolio"

    def test_alle_deutschen_dateien_geladen(self):
        t = _setze_sprache("de")
        geladen = t.lade_texte("de")
        for datei in ("oberflaeche", "diagramme", "bericht", "excel"):
            assert any(k.startswith(f"{datei}.") for k in geladen), datei

    def test_englisch_vorhandener_schluessel(self):
        t = _setze_sprache("en")
        assert t.txt("oberflaeche.nav_neues_projekt") == "New Project"

    def test_englisch_fehlender_schluessel_faellt_auf_deutsch_zurueck(
        self, tmp_path, monkeypatch
    ):
        """Simuliert eine unvollstaendige Uebersetzung ueber ein
        isoliertes, synthetisches locales-Verzeichnis (unabhaengig vom
        tatsaechlichen Vollstaendigkeitsgrad der ausgelieferten
        Sprachdateien): fehlt ein Schluessel in der Zielsprache, greift
        automatisch der deutsche Text."""
        import texte

        (tmp_path / "de").mkdir()
        (tmp_path / "de" / "test.yaml").write_text(
            "a: Deutscher Text A\nb: Deutscher Text B\n", encoding="utf-8"
        )
        (tmp_path / "xx").mkdir()
        (tmp_path / "xx" / "test.yaml").write_text(
            "a: XX Text A\n", encoding="utf-8"  # 'b' bewusst nicht uebersetzt
        )
        monkeypatch.setattr(texte, "LOCALES_DIR", tmp_path)
        monkeypatch.setenv("NOBIS_SPRACHE", "xx")
        texte.lade_texte.cache_clear()
        assert texte.txt("test.a") == "XX Text A"
        assert texte.txt("test.b") == "Deutscher Text B"
        texte.lade_texte.cache_clear()

    def test_alle_ausgelieferten_sprachen_vollstaendig(self):
        """Die vier gepflegten Sprachen (de/en/fr/es) uebersetzen jeden
        deutschen Schluessel vollstaendig - kein Schluessel fehlt und
        keine Platzhalter weichen ab (Qualitaetssicherung der
        Uebersetzungsdateien selbst, unabhaengig vom Fallback-
        Mechanismus)."""
        import re

        import yaml

        from texte import LOCALES_DIR, SPRACHEN

        def platzhalter(text: str) -> set[str]:
            return set(re.findall(r"\{(\w+)\}", text))

        for datei in ("oberflaeche", "diagramme", "bericht", "excel"):
            de = yaml.safe_load((LOCALES_DIR / "de" / f"{datei}.yaml").read_text())
            for code in SPRACHEN:
                pfad = LOCALES_DIR / code / f"{datei}.yaml"
                assert pfad.exists(), pfad
                inhalt = yaml.safe_load(pfad.read_text())
                assert set(inhalt) == set(de), f"{code}/{datei}: Schlüssel weichen ab"
                for schluessel, text_de in de.items():
                    assert platzhalter(text_de) == platzhalter(inhalt[schluessel]), (
                        f"{code}/{datei}.{schluessel}: Platzhalter weichen ab"
                    )


class TestSprachdropdownEndToEnd:
    """End-to-End-Tests des Sprach-Dropdowns oben rechts in der
    laufenden Streamlit-App (streamlit_app.py): Umschalten wirkt sofort
    auf Navigation, Buttons und den PDF-Export."""

    _NAV_ERWARTET = {
        "de": ["Portfolio", "Neues Projekt", "Ausschreibung", "Globale Annahmen"],
        "en": ["Portfolio", "New Project", "Auction", "Global Assumptions"],
        "fr": ["Portefeuille", "Nouveau projet", "Appel d'offres",
               "Hypothèses globales"],
        "es": ["Cartera", "Nuevo proyecto", "Subasta", "Supuestos globales"],
    }

    @pytest.mark.parametrize("code", ["de", "en", "fr", "es"])
    def test_dropdown_uebersetzt_navigation(self, code):
        """Jede Sprache einzeln, mit frischer AppTest-Instanz (mehrfaches
        Umschalten in derselben Session ist eine Einschraenkung des
        Streamlit-Testframeworks selbst, nicht der App - siehe
        docs/AppTest format_func caching)."""
        from streamlit.testing.v1 import AppTest

        at = AppTest.from_file(
            str(Path(__file__).parent.parent / "streamlit_app.py"),
            default_timeout=300,
        )
        at.run()
        assert not at.exception
        if code != "de":
            knopf = [
                b for b in at.button if b.key == f"sprachauswahl_{code}"
            ][0]
            knopf.click()
            at.run()
            assert not at.exception
        assert at.sidebar.radio[0].options == self._NAV_ERWARTET[code]

    def test_pdf_export_folgt_gewaehlter_sprache(self):
        """Der PDF-Bericht wird in der ueber das Dropdown gewaehlten
        Sprache erzeugt - Button-Label, Kapitelueberschrift und
        Kapitel-8-Fliesstext."""
        import io

        from pypdf import PdfReader
        from streamlit.testing.v1 import AppTest

        at = AppTest.from_file(
            str(Path(__file__).parent.parent / "streamlit_app.py"),
            default_timeout=300,
        )
        at.run()
        [b for b in at.button if b.key == "sprachauswahl_en"][0].click()
        at.run()
        assert not at.exception

        [b for b in at.button if b.key and b.key.startswith("open_")][0].click()
        at.run()
        assert not at.exception
        pdf_btn = [b for b in at.button if "PDF" in (b.label or "")][0]
        assert pdf_btn.label == "Create PDF report"
        pdf_btn.click()
        at.run(timeout=300)
        assert not at.exception

        pdf_bytes = next(
            v for k, v in at.session_state.filtered_state.items()
            if k.startswith("pdf_bericht_") and v
        )
        text = "\n".join(
            s.extract_text() for s in PdfReader(io.BytesIO(pdf_bytes)).pages
        )
        assert "Management Summary" in text
        assert "EAG Auction Model" in text
        assert "Austria has been awarding the EAG market premium" in text

    def test_voellig_unbekannter_schluessel_liefert_schluessel(self):
        t = _setze_sprache("de")
        assert t.txt("oberflaeche.gibt_es_nicht") == "oberflaeche.gibt_es_nicht"

    def test_nicht_existierende_sprache_faellt_komplett_auf_deutsch(self):
        t = _setze_sprache("xx")
        assert t.txt("oberflaeche.nav_portfolio") == "Portfolio"


class TestPlatzhalter:
    def test_platzhalter_werden_eingesetzt(self):
        t = _setze_sprache("de")
        ergebnis = t.txt("oberflaeche.portfolio_toggle_inaktive", anzahl=3)
        assert "3" in ergebnis and "{anzahl}" not in ergebnis

    def test_ohne_platzhalter_bleibt_text_unveraendert(self):
        t = _setze_sprache("de")
        # Text ohne {..}-Platzhalter: keine Formatierung noetig, auch
        # wenn geschweifte Klammern (z. B. aus Hovertemplates) vorkaemen.
        assert t.txt("oberflaeche.btn_speichern") == "Speichern"

    def test_fehlender_platzhalter_wirft_nicht(self):
        t = _setze_sprache("de")
        # Fail-soft: falscher/fehlender Platzhaltername darf nicht crashen.
        ergebnis = t.txt("oberflaeche.portfolio_toggle_inaktive", falsch=1)
        assert "{anzahl}" in ergebnis


class TestExcelTexte:
    def test_prefix_wird_entfernt(self):
        t = _setze_sprache("de")
        flach = t.excel_texte()
        assert "blatt_uebersicht" in flach
        assert flach["blatt_uebersicht"] == "Übersicht"
        assert not any(k.startswith("excel.") for k in flach)

    def test_verwendung_in_io_ergebnis_excel(self):
        from engine.io_ergebnis_excel import _t

        _setze_sprache("de")
        assert _t("kpi_nennleistung") == "Nennleistung"
        assert "300" in _t("sektion_monte_carlo", n_mc=300)


class TestBerichtKapitel8Ausgelagert:
    """Die zuvor als bekannte Luecke dokumentierten Fliesstext-Absaetze
    aus PDF-Kapitel 8 sind jetzt vollstaendig in bericht.yaml ausgelagert
    und mit Platzhaltern fuer die eingesetzten Kennzahlen versehen."""

    def test_alle_kapitel8_schluessel_vorhanden(self):
        t = _setze_sprache("de")
        schluessel = (
            "abschnitt_auktion_fitting", "auktion_historie_intro",
            "auktion_historie_unterzeichnet", "auktion_historie_wettbewerb",
            "abb_14_caption", "auktion_fitting_intro", "abb_15_caption",
            "modell_verteilungsfamilie", "modell_kalibrierung",
            "modell_punktprognose_intro", "modell_punktprognose_ergebnis",
            "modell_unsicherheit_intro", "modell_zuschlagsdichte_intro",
            "abb_16_caption", "abb_17_caption",
            "tab_3_spalte_wahrscheinlichkeit", "tab_3_spalte_gebotswert",
            "tab_3_caption",
        )
        for s in schluessel:
            text = t.txt(f"bericht.{s}")
            assert text and text != f"bericht.{s}", s

    def test_platzhalter_werden_korrekt_eingesetzt(self):
        t = _setze_sprache("de")
        text = t.txt("bericht.auktion_historie_intro", n_runden=15,
                     erste_datum="12/2022")
        assert "15 Runden seit 12/2022" in text
        text2 = t.txt("bericht.abb_16_caption", grenzzuschlag_ct="6,65 ct/kWh",
                      projektwert_ct="6,50 ct/kWh")
        assert "6,65 ct/kWh" in text2 and "6,50 ct/kWh" in text2

    def test_formel_zeile_platzhalter(self):
        t = _setze_sprache("de")
        text = t.txt("bericht.modell_punktprognose_ergebnis",
                     formel_zeile="Test-Stützstellen: 1 → 2 ct.")
        assert text.endswith("Test-Stützstellen: 1 → 2 ct.")

    def test_report_nutzt_ausgelagerte_texte(self):
        """Der generierte PDF-Bericht enthaelt die ausgelagerten Absaetze
        unveraendert (End-to-End ueber den bestehenden Berichts-Service)."""
        import io

        from pypdf import PdfReader

        from app import services

        _setze_sprache("de")
        pdf = services.build_project_report("template-agri", 0.08)
        text = "\n".join(
            s.extract_text() for s in PdfReader(io.BytesIO(pdf)).pages
        )
        assert "Österreich vergibt die EAG-Marktprämie" in text
        assert "Seit Juli 2025 ist das Bild gekippt" in text


class TestWeitereSeitenUebersetzt:
    """Stichproben auf Seiten jenseits der Navigation/des PDF-Exports, die
    in einer nachtraeglichen, gruendlicheren Uebersetzungsrunde ergaenzt
    wurden (Formular, Ausschreibungsseite, Globale Annahmen, Sidebar)."""

    @pytest.mark.parametrize(
        "code,nav_code,erwartet",
        [
            ("es", "neu", "Crear nuevo proyecto"),
            ("es", "auktion", "Simulación de subasta EAG"),
            ("fr", "annahmen", "Scénarios de prix de marché"),
            ("en", "neu", "Create new project"),
        ],
    )
    def test_seite_uebersetzt(self, code, nav_code, erwartet, monkeypatch):
        """Sprache ueber NOBIS_SPRACHE statt Dropdown gesetzt: isoliert die
        eigentliche Frage (rendert die Seite korrekt uebersetzt?) von der
        Dropdown-Rerun-Mechanik, die bereits separat in
        TestSprachdropdownEndToEnd geprueft ist. Zwei Runs kurz
        hintereinander (Dropdown+Rerun mitten im Skript) bringen AppTests
        interne Widget-Serialisierung sonst durcheinander (reines
        Testframework-Artefakt, siehe dortige Erklaerung)."""
        from streamlit.testing.v1 import AppTest

        import texte

        monkeypatch.setenv("NOBIS_SPRACHE", code)
        texte.lade_texte.cache_clear()
        at = AppTest.from_file(
            str(Path(__file__).parent.parent / "streamlit_app.py"),
            default_timeout=300,
        )
        at.run()
        assert not at.exception
        at.sidebar.radio[0].set_value(nav_code)
        at.run()
        assert not at.exception, at.exception
        texte_ = " ".join(m.value for m in at.markdown if m.value)
        texte_ += " ".join(s.value for s in at.subheader if s.value)
        if at.expander:
            texte_ += " ".join(e.label or "" for e in at.expander)
        assert erwartet in texte_, f"'{erwartet}' nicht gefunden für {code}/{nav_code}"

    def test_sidebar_uebersetzt(self):
        from streamlit.testing.v1 import AppTest

        at = AppTest.from_file(
            str(Path(__file__).parent.parent / "streamlit_app.py"),
            default_timeout=300,
        )
        at.run()
        assert not at.exception
        [b for b in at.button if b.key == "sprachauswahl_en"][0].click()
        at.run()
        assert not at.exception
        labels = [e.label for e in at.sidebar.expander]
        assert any("Save / restore projects" in (lbl or "") for lbl in labels)
        assert any("Save / restore global assumptions" in (lbl or "") for lbl in labels)


class TestFlaggenIcons:
    """Regressionstest fuer den Wechsel von Emoji-Flaggen (nicht auf
    allen Systemen/Browsern darstellbar, insbesondere Windows) auf
    echte Bild-Icons im Sprach-Popover."""

    def test_vier_flaggendateien_vorhanden_und_gueltig(self):
        from PIL import Image

        from app.config import FLAGS_DIR
        from texte import SPRACHEN

        for eintrag in SPRACHEN.values():
            pfad = FLAGS_DIR / f"{eintrag['flagge']}.png"
            assert pfad.exists(), pfad
            with Image.open(pfad) as bild:
                assert bild.format == "PNG"
                assert bild.width > 0 and bild.height > 0

    def test_popover_zeigt_flaggenbilder_und_schaltet_um(self):
        """5 Bilder im Baum (1 Logo + 4 Flaggen) belegen, dass echte
        Bild-Icons statt Emoji im Popover eingebettet sind; die
        Sprachumschaltung per Button-Klick funktioniert weiterhin."""
        from streamlit.testing.v1 import AppTest

        at = AppTest.from_file(
            str(Path(__file__).parent.parent / "streamlit_app.py"),
            default_timeout=300,
        )
        at.run()
        assert not at.exception
        assert len(list(at.image)) == 5

        sprach_buttons = [
            b for b in at.button if b.key and b.key.startswith("sprachauswahl_")
        ]
        assert len(sprach_buttons) == 4
        assert {b.key for b in sprach_buttons} == {
            "sprachauswahl_de", "sprachauswahl_en",
            "sprachauswahl_fr", "sprachauswahl_es",
        }

        [b for b in at.button if b.key == "sprachauswahl_en"][0].click()
        at.run()
        assert not at.exception
        assert "New Project" in at.sidebar.radio[0].options


class TestRebrandingNobisAnalytics:
    """Regressionstest fuer den Rebrand von Trianel/TEA auf Nobis
    Analytics: neues Logo/Favicon eingebunden, keine 'TEA'-Reste im
    laufenden Programm oder in erzeugten Dokumenten."""

    def test_config_zeigt_neue_marke(self):
        from app.config import APP_TITLE, FAVICON_PATH, LOGO_PATH

        assert APP_TITLE == "Nobis Analytics"
        assert LOGO_PATH.name == "nobis_logo.png"
        assert LOGO_PATH.exists()
        assert FAVICON_PATH.exists()

    def test_kein_tea_in_sprachdateien(self):
        import pathlib

        for pfad in pathlib.Path("locales").rglob("*.yaml"):
            inhalt = pfad.read_text(encoding="utf-8")
            assert "TEA" not in inhalt, pfad

    def test_app_header_zeigt_logo_ohne_tea(self):
        from streamlit.testing.v1 import AppTest

        at = AppTest.from_file(
            str(Path(__file__).parent.parent / "streamlit_app.py"),
            default_timeout=300,
        )
        at.run()
        assert not at.exception
        md = " ".join(m.value for m in at.markdown if m.value)
        assert "TEA" not in md
        assert len(list(at.image)) == 5  # Logo + 4 Flaggen

    def test_pdf_bericht_ohne_tea(self):
        import io

        from pypdf import PdfReader

        from app import services

        pdf = services.build_project_report("template-agri", 0.08)
        text = "\n".join(s.extract_text() for s in PdfReader(io.BytesIO(pdf)).pages)
        assert "TEA" not in text
        assert "Nobis Analytics" in text

    def test_excel_export_ohne_tea(self):
        import io

        from openpyxl import load_workbook

        from app import services

        daten = services.build_pipeline_excel(n_mc=30)
        wb = load_workbook(io.BytesIO(daten))
        titel = wb["Übersicht"]["A1"].value
        assert "TEA" not in titel
        assert "Nobis Analytics" in titel


class TestFarbschemaNobisAnalytics:
    """Regressionstest fuer den Wechsel von Trianel-Rot/Tannengruen auf
    die aus dem Nobis-Analytics-Logo abgeleitete Tuerkis/Navy-Palette:
    keine alten Markenfarben mehr in Colors-Klasse, PDF-Report-Modul
    oder erzeugten Dokumenten."""

    _ALTE_FARBEN = {"#BE172B", "#143530", "#2E5A52", "#5B6B66",
                    "#8AA6A0", "#E1E8E5", "#F6F9F8", "#A31425"}

    def test_colors_klasse_ohne_trianel_farben(self):
        from app.theme import Colors

        werte = [Colors.BRAND, Colors.INK, Colors.INK_SOFT, Colors.MUTED,
                 Colors.NEUTRAL, Colors.LINE, Colors.WASH, *Colors.SERIES]
        assert not (set(werte) & self._ALTE_FARBEN)
        assert Colors.BRAND == "#167B88"
        assert Colors.INK == "#14304F"

    def test_report_modul_ohne_trianel_farben(self):
        import app.report as report

        werte = {report.BRAND, report.INK, report.INK_SOFT, report.MUTED,
                 report.NEUTRAL, report.LINE, report.WASH, *report.SERIES}
        assert not (werte & self._ALTE_FARBEN)
        # App-Theme und PDF-Modul muessen exakt dieselbe Palette fahren.
        from app.theme import Colors
        assert report.BRAND == Colors.BRAND
        assert report.INK == Colors.INK

    def test_keine_alten_farbcodes_im_quellcode(self):
        """app/branding.py haelt BEIDE Paletten bewusst vor (Registry
        des verdeckten Marken-Schalters, siehe TestVerdeckterMarkenSchalter)
        und wird deshalb ausgenommen - ueberall sonst duerfen die alten
        Trianel-Farbcodes nicht als aktive/Standardwerte auftauchen."""
        import pathlib

        muster = self._ALTE_FARBEN | {"20, 53, 48", "20,53,48",
                                      "138, 166, 160", "138,166,160"}
        for pfad in pathlib.Path("app").rglob("*.py"):
            if pfad.name == "branding.py":
                continue
            inhalt = pfad.read_text(encoding="utf-8")
            for alt in muster:
                assert alt not in inhalt, f"{pfad}: {alt}"

    def test_pdf_bericht_ohne_rote_akzentfarbe(self):
        """Rendert den Bericht und prueft die Pixelfarben: keine
        Trianel-Rot-Pixel (Deckblatt-Linie, Kapitelnummern), stattdessen
        das neue Tuerkis vorhanden. Erfordert poppler (pdftoppm); wird
        uebersprungen, falls nicht installiert."""
        import shutil
        import subprocess
        import tempfile

        if not shutil.which("pdftoppm"):
            pytest.skip("pdftoppm (poppler) nicht installiert")

        import pathlib

        import numpy as np
        from PIL import Image

        from app import services

        pdf = services.build_project_report("template-agri", 0.08)
        with tempfile.TemporaryDirectory() as tmp:
            pdf_pfad = f"{tmp}/bericht.pdf"
            with open(pdf_pfad, "wb") as f:
                f.write(pdf)
            subprocess.run(
                ["pdftoppm", "-png", "-r", "80", "-f", "1", "-l", "1",
                 pdf_pfad, f"{tmp}/seite"],
                check=True,
            )
            bild = Image.open(next(pathlib.Path(tmp).glob("seite-*.png"))).convert("RGB")
            arr = np.array(bild).reshape(-1, 3).astype(int)

            def pixel_nahe(farbe, toleranz=18):
                dist = np.sqrt(((arr - np.array(farbe)) ** 2).sum(axis=1))
                return int((dist < toleranz).sum())

            assert pixel_nahe((190, 23, 43)) == 0, "Trianel-Rot noch im PDF sichtbar"
            assert pixel_nahe((22, 123, 136)) > 100, "Neues Tuerkis nicht gefunden"


class TestKopfzeileLaeuftSchnellDurch:
    """Regressionstest fuer einen Einrueckungsfehler: st.rerun() im
    Sprachauswahl-Popover lief zwischenzeitlich bei JEDER
    Schleifeniteration ueber die vier Sprachen (nicht nur bei Klick),
    was einen Rerun-Endlosloop und dadurch u.a. eine visuell
    abgeschnittene Kopfzeile verursachte (Logo/Popover kamen nie zum
    fertigen Rendern)."""

    def test_app_rendert_schnell_ohne_endlosschleife(self):
        import time

        from streamlit.testing.v1 import AppTest

        start = time.monotonic()
        at = AppTest.from_file(
            str(Path(__file__).parent.parent / "streamlit_app.py"),
            default_timeout=30,
        )
        at.run()
        dauer = time.monotonic() - start
        assert not at.exception, at.exception
        assert dauer < 15, (
            f"Initialer Render dauerte {dauer:.1f}s - deutet auf einen "
            "Rerun-Loop hin (z.B. unbedingtes st.rerun() in einer Schleife)"
        )

    def test_kein_unbedingtes_rerun_in_sprachschleife(self):
        """Statische Absicherung: st.rerun() darf im Sprachauswahl-Block
        nur INNERHALB des Klick-if stehen, nicht auf Schleifenebene."""
        quelle = Path("streamlit_app.py").read_text(encoding="utf-8")
        block = quelle[quelle.index("for code in _sprachcodes:"):]
        block = block[:block.index("st.markdown('<div class=\"app-header-rule\"")]
        zeilen = block.splitlines()
        rerun_zeilen = [z for z in zeilen if "st.rerun()" in z]
        assert len(rerun_zeilen) == 1
        einrueckung_rerun = len(rerun_zeilen[0]) - len(rerun_zeilen[0].lstrip())
        einrueckung_session = next(
            len(z) - len(z.lstrip()) for z in zeilen if "SESSION_KEY] = code" in z
        )
        assert einrueckung_rerun == einrueckung_session, (
            "st.rerun() ist nicht auf derselben Einrueckungsebene wie die "
            "Session-State-Zuweisung - liegt es ausserhalb des if-Blocks, "
            "feuert es bei jeder Schleifeniteration unbedingt."
        )


class TestVerdeckterMarkenSchalter:
    """End-to-End-Tests des verdeckten Marken-Schalters (app/branding.py):
    URL-Parameter ?marke=trianel zeigt die vorherige Trianel-Gestaltung
    (Original-Assets, wiederhergestellt aus einem frueheren Archiv),
    ohne Parameter gilt Nobis Analytics."""

    def test_registry_haelt_beide_marken_vollstaendig(self):
        from app.branding import MARKEN

        assert set(MARKEN) == {"nobis", "trianel"}
        for code, marke in MARKEN.items():
            for schluessel in ("app_titel", "kopfzeile_titel", "logo",
                              "logo_breite", "favicon", "farben"):
                assert schluessel in marke, f"{code}.{schluessel}"
            for farbschluessel in ("BRAND", "INK", "INK_SOFT", "MUTED",
                                   "NEUTRAL", "LINE", "WASH"):
                assert farbschluessel in marke["farben"], f"{code}.{farbschluessel}"

    def test_trianel_assets_sind_die_wiederhergestellten_originale(self):
        """Keine Neuerstellung/Annaeherung - die Original-Bilddateien
        aus dem Archiv vor dem Rebrand."""
        from app.branding import MARKEN

        logo = MARKEN["trianel"]["logo"]
        favicon = MARKEN["trianel"]["favicon"]
        assert logo.exists() and logo.stat().st_size > 10_000
        assert favicon.exists() and favicon.stat().st_size > 500

    def test_ohne_parameter_zeigt_nobis_analytics(self):
        from streamlit.testing.v1 import AppTest

        at = AppTest.from_file(
            str(Path(__file__).parent.parent / "streamlit_app.py"),
            default_timeout=60,
        )
        at.run()
        assert not at.exception, at.exception
        md = " ".join(m.value for m in at.markdown if m.value)
        assert "PV-Projektbewertung" in md
        assert "TEA PV-Projektbewertung" not in md

    def test_parameter_trianel_zeigt_alte_gestaltung(self):
        from streamlit.testing.v1 import AppTest

        try:
            at = AppTest.from_file(
                str(Path(__file__).parent.parent / "streamlit_app.py"),
                default_timeout=60,
            )
            at.query_params["marke"] = "trianel"
            at.run()
            assert not at.exception, at.exception
            md = " ".join(m.value for m in at.markdown if m.value)
            assert "TEA PV-Projektbewertung" in md
            # Farben tatsaechlich umgeschaltet (nicht nur der Text).
            from app.theme import Colors
            assert Colors.BRAND == "#BE172B"
            assert Colors.INK == "#143530"
        finally:
            # Colors ist bewusst einfacher, prozessweiter Zustand (siehe
            # Einschraenkung in app/branding.py) - ohne Reset wuerde
            # dieser Test alle NACHFOLGENDEN Tests in der gesamten
            # Suite verunreinigen, nicht nur diese Datei.
            from app.branding import MARKEN
            from app.theme import wende_farben_an
            wende_farben_an(MARKEN["nobis"]["farben"])

    def test_unbekannter_parameter_faellt_auf_nobis_zurueck(self):
        from streamlit.testing.v1 import AppTest

        at = AppTest.from_file(
            str(Path(__file__).parent.parent / "streamlit_app.py"),
            default_timeout=60,
        )
        at.query_params["marke"] = "does-not-exist"
        at.run()
        assert not at.exception, at.exception
        md = " ".join(m.value for m in at.markdown if m.value)
        assert "TEA PV-Projektbewertung" not in md

    def test_pdf_bericht_folgt_dem_schalter(self):
        """Der PDF-Export respektiert die aktive Marke (Farben, Logo,
        Fusszeilen-/Autor-Signatur), nicht nur die Live-App.
        services.build_project_report() ermittelt die Marke intern
        selbst ueber aktive_marke() - fuer den Test ausserhalb einer
        echten Streamlit-Session (dort faellt aktive_marke() mangels
        Session-Kontext auf "nobis" zurueck) wird sie deshalb gemockt,
        nicht nur die Farben vorbelegt."""
        import io
        from unittest.mock import patch

        from pypdf import PdfReader

        from app import branding, report, services

        try:
            with patch(
                "app.branding.aktive_marke",
                return_value=branding.MARKEN["trianel"],
            ):
                pdf = services.build_project_report("template-agri", 0.08)
            assert pdf is not None
            text = "\n".join(
                s.extract_text() for s in PdfReader(io.BytesIO(pdf)).pages
            )
            assert "Nobis Analytics" not in text
            assert "TEA PV-Projektbewertung" in text
        finally:
            # Colors/report sind bewusst einfacher, prozessweiter Zustand
            # (siehe Einschraenkung in app/branding.py) - ohne Reset
            # wuerde dieser Test alle NACHFOLGENDEN Tests in der
            # gesamten Suite verunreinigen, nicht nur diese Datei.
            report.wende_farben_an(branding.MARKEN["nobis"]["farben"])

    def test_nach_trianel_test_faellt_theme_auf_nobis_zurueck(self):
        """Reihenfolge-Unabhaengigkeit: stellt sicher, dass ein
        vorheriger Test dieser Klasse den globalen Colors-Zustand nicht
        fuer nachfolgende Tests (in DIESER oder anderen Testdateien)
        verschmutzt zuruecklaesst."""
        from streamlit.testing.v1 import AppTest

        at = AppTest.from_file(
            str(Path(__file__).parent.parent / "streamlit_app.py"),
            default_timeout=60,
        )
        at.run()
        assert not at.exception
        from app.theme import Colors
        assert Colors.BRAND == "#167B88"

    def test_excel_export_folgt_dem_schalter(self):
        """Auch der Pipeline-Excel-Export (io_ergebnis_excel.py, reine
        Engine-Funktion ohne Streamlit-Abhaengigkeit) uebernimmt den
        Markennamen als expliziten Parameter statt festen Text."""
        import io

        from openpyxl import load_workbook

        from engine import pipeline_ergebnis_excel
        from tests.conftest import _baue_global_assumptions, _baue_projekt

        projekt = _baue_projekt()
        daten = pipeline_ergebnis_excel(
            [(projekt, projekt.name)], _baue_global_assumptions(),
            n_mc=30, marken_name="TEA PV-Projektbewertung",
        )
        wb = load_workbook(io.BytesIO(daten))
        titel = wb["Übersicht"]["A1"].value
        assert "TEA PV-Projektbewertung" in titel
        assert "Nobis Analytics" not in titel
