"""Tests fuer Zeitachse, Energieproduktion und Betriebskosten."""

from __future__ import annotations

from datetime import date

import numpy as np
import pytest

from engine import OpexItem
from engine.energy import calculate_energy_production
from engine.opex import calculate_opex
from engine.pipeline import resolve_assumptions
from engine.timeline import build_timeline


class TestTimeline:
    def test_jahresstart_januar_hat_volle_prorata(self):
        timeline = build_timeline(date(2027, 1, 1), 3)
        assert len(timeline) == 3
        assert timeline["pro_rata_faktor"].iloc[0] == pytest.approx(1.0)
        assert bool(timeline["ist_letztes_jahr"].iloc[-1]) is True

    def test_unterjaehriger_start_hat_anteiliges_erstes_jahr(self):
        # Start 1. Juli -> zweites Halbjahr = 184 Tage.
        timeline = build_timeline(date(2027, 7, 1), 2)
        assert timeline["pro_rata_faktor"].iloc[0] == pytest.approx(184 / 365)
        assert timeline["pro_rata_faktor"].iloc[1] == pytest.approx(1.0)

    def test_laufzeit_null_wird_abgelehnt(self):
        with pytest.raises(ValueError):
            build_timeline(date(2027, 1, 1), 0)


class TestZinsmethode:
    """Zinsberechnungsmethode fuer das (moeglicherweise unterjaehrige)
    erste Betriebsjahr: deutsche (30/360) vs. oesterreichische
    (act/365) Konvention - siehe engine.timeline.erstjahr_zins_pro_rata
    und engine.financing.calculate_financing."""

    def test_januar_start_beide_methoden_voller_faktor(self):
        from engine.models import ZinsMethode
        from engine.timeline import erstjahr_zins_pro_rata

        for methode in (ZinsMethode.OESTERREICH, ZinsMethode.DEUTSCH):
            assert erstjahr_zins_pro_rata(date(2028, 1, 1), methode) == pytest.approx(1.0)

    def test_oesterreich_deckt_sich_mit_produktions_pro_rata(self):
        """Dieselbe taggenaue Logik wie build_timeline()'s
        pro_rata_faktor fuer das erste Jahr - keine zweite,
        abweichende Zeitachse fuer die Finanzierung."""
        from engine.models import ZinsMethode
        from engine.timeline import build_timeline, erstjahr_zins_pro_rata

        for monat in (3, 6, 9, 12):
            start = date(2028, monat, 1)
            timeline = build_timeline(start, 2)
            faktor_timeline = timeline["pro_rata_faktor"].iloc[0]
            faktor_zins = erstjahr_zins_pro_rata(start, ZinsMethode.OESTERREICH)
            assert faktor_zins == pytest.approx(faktor_timeline)

    def test_deutsch_30_360_rechenformel(self):
        from engine.models import ZinsMethode
        from engine.timeline import erstjahr_zins_pro_rata

        # Juni: Restmonate inkl. Juni = 7 -> 7*30/360
        assert erstjahr_zins_pro_rata(
            date(2028, 6, 15), ZinsMethode.DEUTSCH
        ) == pytest.approx(7 * 30 / 360)
        # Dezember: nur 1 Restmonat -> 30/360
        assert erstjahr_zins_pro_rata(
            date(2028, 12, 1), ZinsMethode.DEUTSCH
        ) == pytest.approx(30 / 360)
        # 30/360 ignoriert den Tag im Monat (kaufmaennische Konvention).
        assert erstjahr_zins_pro_rata(
            date(2028, 6, 1), ZinsMethode.DEUTSCH
        ) == erstjahr_zins_pro_rata(date(2028, 6, 30), ZinsMethode.DEUTSCH)

    def test_deutsch_und_oesterreich_nah_beieinander_aber_nicht_gleich(self):
        from engine.models import ZinsMethode
        from engine.timeline import erstjahr_zins_pro_rata

        start = date(2028, 6, 1)
        at = erstjahr_zins_pro_rata(start, ZinsMethode.OESTERREICH)
        de = erstjahr_zins_pro_rata(start, ZinsMethode.DEUTSCH)
        assert at != de
        assert abs(at - de) < 0.01

    def test_financing_wendet_faktor_nur_auf_erstes_jahr_an(self):
        from engine.financing import calculate_financing
        from engine.models import TilgungsArt
        from engine.timeline import build_timeline

        timeline = build_timeline(date(2028, 7, 1), 3)
        voll = calculate_financing(
            timeline, 1_000_000, 0.2, 0.05, 3, TilgungsArt.ANNUITAET,
            erstjahr_zins_faktor=1.0,
        )
        halb = calculate_financing(
            timeline, 1_000_000, 0.2, 0.05, 3, TilgungsArt.ANNUITAET,
            erstjahr_zins_faktor=0.5,
        )
        # Jahr 1: Zinsen exakt halbiert.
        assert halb["zinsen_eur"].iloc[0] == pytest.approx(
            voll["zinsen_eur"].iloc[0] * 0.5
        )
        # Folgejahre (2, 3): NICHT vom Faktor betroffen (nur jahr==1
        # wird reduziert) - beide Reihen duerfen ab Jahr 2 aber wegen
        # der unterschiedlichen Tilgungshoehe in Jahr 1 divergieren,
        # der Faktor selbst wird aber nicht mehr angewendet:
        assert halb["darlehensstand_bop_eur"].iloc[0] == pytest.approx(
            voll["darlehensstand_bop_eur"].iloc[0]
        )

    def test_pipeline_reduziert_zinsen_bei_unterjaehrigem_start(
        self, project, global_assumptions
    ):
        """End-to-End: ein Projekt mit Inbetriebnahme im Juni zahlt im
        ersten Betriebsjahr deutlich weniger Zinsen als eines mit
        Inbetriebnahme im Januar (identische sonstige Annahmen) - vor
        dieser Aenderung wurde immer ein volles Jahr Zinsen berechnet,
        unabhaengig vom Inbetriebnahmemonat."""
        from engine.pipeline import run_valuation

        project.inbetriebnahme_monat = 1
        r_januar = run_valuation(project, global_assumptions)
        zinsen_januar = r_januar.cashflow.data.query("jahr == 1")["zinsen_eur"].iloc[0]

        project.inbetriebnahme_monat = 6
        r_juni = run_valuation(project, global_assumptions)
        zinsen_juni = r_juni.cashflow.data.query("jahr == 1")["zinsen_eur"].iloc[0]

        assert zinsen_juni < zinsen_januar
        # Grobe Erwartung: knapp die Haelfte (Juni -> ca. 7/12 Jahr).
        assert 0.5 < zinsen_juni / zinsen_januar < 0.65

    def test_deutsch_vs_oesterreich_wirkt_sich_auf_irr_aus(
        self, project, global_assumptions
    ):
        from engine.models import ZinsMethode
        from engine.pipeline import run_valuation

        project.inbetriebnahme_monat = 6
        global_assumptions.zinsmethode = ZinsMethode.OESTERREICH
        r_at = run_valuation(project, global_assumptions)
        global_assumptions.zinsmethode = ZinsMethode.DEUTSCH
        r_de = run_valuation(project, global_assumptions)

        assert r_at.kpis.equity_irr != r_de.kpis.equity_irr
        # Die Methoden liegen dicht beieinander - kein grosser Sprung.
        assert abs(r_at.kpis.equity_irr - r_de.kpis.equity_irr) < 0.01

    def test_januar_start_beide_methoden_identisches_ergebnis(
        self, project, global_assumptions
    ):
        """Fuer volle Kalenderjahre (Inbetriebnahme im Januar) darf die
        Methodenwahl keinen Unterschied machen."""
        from engine.models import ZinsMethode
        from engine.pipeline import run_valuation

        project.inbetriebnahme_monat = 1
        global_assumptions.zinsmethode = ZinsMethode.OESTERREICH
        r_at = run_valuation(project, global_assumptions)
        global_assumptions.zinsmethode = ZinsMethode.DEUTSCH
        r_de = run_valuation(project, global_assumptions)

        assert r_at.kpis.equity_irr == pytest.approx(r_de.kpis.equity_irr)

    def test_yaml_roundtrip(self, tmp_path):
        from engine.io_yaml import (
            load_global_assumptions_yaml,
            save_global_assumptions_yaml,
        )
        from engine.models import ZinsMethode

        ga = load_global_assumptions_yaml("data/global_assumptions.yaml")
        ga.zinsmethode = ZinsMethode.DEUTSCH
        pfad = tmp_path / "ga.yaml"
        save_global_assumptions_yaml(ga, str(pfad))
        ga2 = load_global_assumptions_yaml(str(pfad))
        assert ga2.zinsmethode == ZinsMethode.DEUTSCH

    def test_excel_roundtrip(self):
        from engine.io_excel import (
            excel_to_global_assumptions,
            global_assumptions_to_excel,
        )
        from engine.io_yaml import load_global_assumptions_yaml
        from engine.models import ZinsMethode

        ga = load_global_assumptions_yaml("data/global_assumptions.yaml")
        ga.zinsmethode = ZinsMethode.DEUTSCH
        ga2 = excel_to_global_assumptions(global_assumptions_to_excel(ga))
        assert ga2.zinsmethode == ZinsMethode.DEUTSCH

    def test_excel_ohne_zinsmethode_spalte_faellt_auf_oesterreich_zurueck(self):
        """Rueckwaertskompatibilitaet: aeltere exportierte Excel-Dateien
        ohne die neue Zeile duerfen beim Import nicht scheitern."""
        import io

        from openpyxl import load_workbook

        from engine.io_excel import (
            excel_to_global_assumptions,
            global_assumptions_to_excel,
        )
        from engine.io_yaml import load_global_assumptions_yaml
        from engine.models import ZinsMethode

        ga = load_global_assumptions_yaml("data/global_assumptions.yaml")
        xl = global_assumptions_to_excel(ga)
        wb = load_workbook(io.BytesIO(xl))
        ws = wb["Einstellungen"]
        for row in ws.iter_rows():
            if row[0].value == "zinsmethode":
                ws.delete_rows(row[0].row)
                break
        puffer = io.BytesIO()
        wb.save(puffer)
        ga2 = excel_to_global_assumptions(puffer.getvalue())
        assert ga2.zinsmethode == ZinsMethode.OESTERREICH


class TestEnergy:
    def test_produktion_ohne_degradation(self, project, global_assumptions):
        assumptions = resolve_assumptions(project, global_assumptions)
        timeline = build_timeline(date(2027, 1, 1), 3)
        energy = calculate_energy_production(timeline, assumptions)
        # 1.000 kWp * 1.000 kWh/kWp = 1 GWh in jedem vollen Jahr.
        assert energy["produktion_kwh"].tolist() == pytest.approx([1e6, 1e6, 1e6])

    def test_degradation_wirkt_ab_jahr_zwei(self, project, global_assumptions):
        global_assumptions.degradation_pct_pa = 0.005
        assumptions = resolve_assumptions(project, global_assumptions)
        timeline = build_timeline(date(2027, 1, 1), 3)
        energy = calculate_energy_production(timeline, assumptions)
        assert energy["produktion_kwh"].iloc[0] == pytest.approx(1e6)
        assert energy["produktion_kwh"].iloc[1] == pytest.approx(1e6 * 0.995)
        assert energy["produktion_kwh"].iloc[2] == pytest.approx(1e6 * 0.995**2)


class TestOpex:
    def test_indexierung_startet_ab_konfiguriertem_jahr(self):
        timeline = build_timeline(date(2027, 1, 1), 3)
        import pandas as pd

        energy = pd.DataFrame({"jahr": [1, 2, 3], "produktion_kwh": [0.0, 0.0, 0.0]})
        items = [
            OpexItem(
                name="Wartung", basiswert_eur_kwp=2.0,
                index_pct_pa=0.02, indexierung_ab_jahr=2,
            )
        ]
        opex = calculate_opex(timeline, items, 1000.0, energy)
        # Jahr 1 und 2: Basis 2.000 €; ab Jahr 3 ein Indexschritt.
        assert opex["Wartung"].iloc[0] == pytest.approx(2000.0)
        assert opex["Wartung"].iloc[1] == pytest.approx(2000.0)
        assert opex["Wartung"].iloc[2] == pytest.approx(2000.0 * 1.02)

    def test_gleichnamige_positionen_werden_addiert(self):
        timeline = build_timeline(date(2027, 1, 1), 1)
        import pandas as pd

        energy = pd.DataFrame({"jahr": [1], "produktion_kwh": [0.0]})
        items = [
            OpexItem(name="Pacht", basiswert_eur_kwp=1.0),
            OpexItem(name="Pacht", basiswert_eur_kwp=2.0),
        ]
        opex = calculate_opex(timeline, items, 1000.0, energy)
        assert opex["Pacht"].iloc[0] == pytest.approx(3000.0)
        # Nur EINE Spalte je Bezeichnung (eindeutiger Legendeneintrag).
        assert list(opex.columns).count("Pacht") == 1

    def test_produktionsbasierte_abgaben(self):
        timeline = build_timeline(date(2027, 1, 1), 1)
        import pandas as pd

        energy = pd.DataFrame({"jahr": [1], "produktion_kwh": [1e6]})
        opex = calculate_opex(
            timeline, [], 1000.0, energy,
            gemeindeabgabe_eur_kwh=0.002,
            direktvermarktungskosten_eur_kwh=0.001,
        )
        assert opex["gemeindeabgabe_eur"].iloc[0] == pytest.approx(2000.0)
        assert opex["direktvermarktungskosten_eur"].iloc[0] == pytest.approx(1000.0)
        assert opex["opex_gesamt_eur"].iloc[0] == pytest.approx(3000.0)


class TestPachtModus:
    """Umsatzbeteiligungs-Pacht mit Mindestpacht (PachtModus): der
    Verpaechter erhaelt MAX(Umsatz x Prozentsatz, Mindestpacht/ha x
    Flaeche, mit Kosteninflation indexiert)."""

    def test_fix_modus_unveraendert_wie_bisher(self):
        """PachtModus.FIX reproduziert exakt das alte Verhalten (Pacht
        als eigene, mit Kosteninflation indexierte Position, ab Jahr 1
        ohne Verzoegerung - wie zuvor ueber die generische
        opex_items-Schleife)."""
        import numpy as np
        import pandas as pd

        from engine.models import PachtModus

        timeline = build_timeline(date(2027, 1, 1), 3)
        energy = pd.DataFrame({"jahr": [1, 2, 3], "produktion_kwh": [0.0] * 3})
        opex = calculate_opex(
            timeline, [], 1000.0, energy,
            kosten_inflation_pct_pa=0.02,
            pacht_modus=PachtModus.FIX, pacht_eur_kwp_jahr=5.0,
        )
        basis = 5.0 * 1000.0
        assert opex["Pacht"].to_numpy() == pytest.approx(
            basis * np.array([1.0, 1.02, 1.02**2])
        )

    def test_umsatzbeteiligung_greift_bei_hohem_umsatz(self):
        """Wenn die Umsatzbeteiligung ueber der Mindestpacht liegt,
        wird die Umsatzbeteiligung gezahlt."""
        import pandas as pd

        from engine.models import PachtModus

        timeline = build_timeline(date(2027, 1, 1), 1)
        energy = pd.DataFrame({"jahr": [1], "produktion_kwh": [0.0]})
        opex = calculate_opex(
            timeline, [], 1000.0, energy,
            pacht_modus=PachtModus.UMSATZBETEILIGUNG,
            pacht_umsatzbeteiligung_pct=0.055,
            pacht_mindestpacht_eur_ha_jahr=1000.0,
            projektflaeche_ha=5.0,
            erloes_eur=pd.Series([500_000.0]).to_numpy(),
        )
        # Umsatzbeteiligung: 500.000 * 5,5% = 27.500 > Mindestpacht 5.000
        assert opex["Pacht"].iloc[0] == pytest.approx(27_500.0)

    def test_mindestpacht_greift_bei_niedrigem_umsatz(self):
        """Wenn die Umsatzbeteiligung unter der Mindestpacht liegt,
        greift die Mindestpacht (Kernanforderung der Funktion)."""
        import pandas as pd

        from engine.models import PachtModus

        timeline = build_timeline(date(2027, 1, 1), 1)
        energy = pd.DataFrame({"jahr": [1], "produktion_kwh": [0.0]})
        opex = calculate_opex(
            timeline, [], 1000.0, energy,
            pacht_modus=PachtModus.UMSATZBETEILIGUNG,
            pacht_umsatzbeteiligung_pct=0.055,
            pacht_mindestpacht_eur_ha_jahr=1000.0,
            projektflaeche_ha=5.0,
            erloes_eur=pd.Series([10_000.0]).to_numpy(),
        )
        # Umsatzbeteiligung: 10.000 * 5,5% = 550 < Mindestpacht 5.000
        assert opex["Pacht"].iloc[0] == pytest.approx(5_000.0)

    def test_mindestpacht_wird_mit_kosteninflation_indexiert(self):
        """Die Mindestpacht steigt Jahr fuer Jahr mit der allgemeinen
        Kosteninflation - genau der vom Nutzer beschriebene Effekt,
        durch den sie in spaeteren Jahren die (nicht automatisch
        mitwachsende) Umsatzbeteiligung uebersteigen kann."""
        import numpy as np
        import pandas as pd

        from engine.models import PachtModus

        timeline = build_timeline(date(2027, 1, 1), 5)
        energy = pd.DataFrame({"jahr": range(1, 6), "produktion_kwh": [0.0] * 5})
        # Umsatz bleibt konstant niedrig -> Mindestpacht dominiert durchgehend,
        # ihr Anstieg ueber die Jahre ist direkt beobachtbar.
        opex = calculate_opex(
            timeline, [], 1000.0, energy,
            kosten_inflation_pct_pa=0.02,
            pacht_modus=PachtModus.UMSATZBETEILIGUNG,
            pacht_umsatzbeteiligung_pct=0.055,
            pacht_mindestpacht_eur_ha_jahr=1000.0,
            projektflaeche_ha=5.0,
            erloes_eur=np.full(5, 1000.0),
        )
        erwartet = 5000.0 * (1.02 ** np.arange(5))
        assert opex["Pacht"].to_numpy() == pytest.approx(erwartet)
        assert opex["Pacht"].is_monotonic_increasing

    def test_uebergang_von_umsatzbeteiligung_zu_mindestpacht_e2e(
        self, project, global_assumptions
    ):
        """End-to-End ueber die volle Pipeline: Pacht = MAX(Umsatz-
        beteiligung, indexierte Mindestpacht) in JEDEM Jahr - und beide
        Seiten gewinnen mindestens einmal ueber die Laufzeit (sonst
        waere die Mindestpacht in diesem Testszenario wirkungslos)."""
        from engine import run_valuation
        from engine.models import PachtModus

        project.pacht_modus = PachtModus.UMSATZBETEILIGUNG
        project.pacht_umsatzbeteiligung_pct = 0.055
        project.pacht_mindestpacht_eur_ha_jahr = 500.0
        project.projektflaeche_ha = 5.0

        r = run_valuation(project, global_assumptions)
        df = r.cashflow.data.query("jahr >= 1")

        jahre_seit_start = (df["jahr"] - 1).clip(lower=0)
        inflation = (1 + global_assumptions.kosten_inflation_pct_pa) ** jahre_seit_start
        umsatzbeteiligung = df["erloes_eur"] * 0.055
        mindestpacht = 500.0 * 5.0 * inflation
        erwartet = umsatzbeteiligung.combine(mindestpacht, max)

        assert df["Pacht"].to_numpy() == pytest.approx(erwartet.to_numpy())
        # Beide Regime kommen im Testzeitraum tatsaechlich vor - sonst
        # waere der MAX()-Vergleich hier nicht aussagekraeftig geprueft.
        assert (umsatzbeteiligung > mindestpacht).any(), (
            "Umsatzbeteiligung gewinnt in keinem Jahr - Testannahmen pruefen"
        )
        assert (mindestpacht > umsatzbeteiligung).any(), (
            "Mindestpacht gewinnt in keinem Jahr - Testannahmen pruefen"
        )

    def test_ohne_flaeche_wirkt_mindestpacht_wie_null(self):
        """Fehlt projektflaeche_ha (None), darf die Mindestpacht nicht
        crashen, sondern wirkt wie 0 - die Umsatzbeteiligung greift
        dann faktisch immer."""
        import pandas as pd

        from engine.models import PachtModus

        timeline = build_timeline(date(2027, 1, 1), 1)
        energy = pd.DataFrame({"jahr": [1], "produktion_kwh": [0.0]})
        opex = calculate_opex(
            timeline, [], 1000.0, energy,
            pacht_modus=PachtModus.UMSATZBETEILIGUNG,
            pacht_umsatzbeteiligung_pct=0.055,
            pacht_mindestpacht_eur_ha_jahr=5000.0,
            projektflaeche_ha=None,
            erloes_eur=pd.Series([100.0]).to_numpy(),
        )
        assert opex["Pacht"].iloc[0] == pytest.approx(100.0 * 0.055)

    def test_gleichnamige_standard_position_wird_addiert_nicht_dupliziert(self):
        """Regressionstest: falls eine globale Standard-OPEX-Position
        zufaellig auch 'Pacht' heisst, darf keine doppelte Spalte
        entstehen (fruehere Fassung dieser Aenderung brach hier mit
        einem Pandas-ValueError)."""
        import pandas as pd

        from engine.models import PachtModus

        timeline = build_timeline(date(2027, 1, 1), 1)
        energy = pd.DataFrame({"jahr": [1], "produktion_kwh": [0.0]})
        kollidierendes_item = OpexItem(name="Pacht", basiswert_eur_kwp=2.0)
        opex = calculate_opex(
            timeline, [kollidierendes_item], 1000.0, energy,
            pacht_modus=PachtModus.FIX, pacht_eur_kwp_jahr=5.0,
        )
        assert list(opex.columns).count("Pacht") == 1
        # 2.0 (Standardposition) + 5.0 (Pachtmodus FIX) = 7.0 EUR/kWp
        assert opex["Pacht"].iloc[0] == pytest.approx(7.0 * 1000.0)
        assert opex["opex_gesamt_eur"].iloc[0] == pytest.approx(7.0 * 1000.0)

    def test_projekt_excel_ohne_pacht_umsatzbeteiligungs_spalten_importierbar(self):
        """Regressionstest (vom Nutzer gemeldet): ein vor v4.19
        exportierter Projekt-Dump kennt die drei neuen Umsatzbeteiligungs-
        Pacht-Spalten noch nicht. Eine fruehere Fassung dieser Aenderung
        hatte zwar die noetige Fallback-Logik pro Feld eingebaut, aber
        eine VORGESCHALTETE strikte Spaltenpruefung liess den Import
        trotzdem mit ValueError scheitern, bevor die Fallback-Logik je
        zum Zug kam."""
        import io

        from openpyxl import load_workbook

        from engine.io_excel import excel_to_projects, projects_to_excel
        from engine.io_yaml import load_project_yaml
        from engine.models import PachtModus

        projekt = load_project_yaml("data/projects/template-agri.yaml")
        xl = projects_to_excel([projekt])
        wb = load_workbook(io.BytesIO(xl))
        ws = wb["Projekte"]
        header = [c.value for c in ws[1]]
        for spalte in (
            "pacht_modus", "pacht_umsatzbeteiligung_pct",
            "pacht_mindestpacht_eur_ha_jahr",
        ):
            ws.delete_cols(header.index(spalte) + 1)
            header.remove(spalte)
        puffer = io.BytesIO()
        wb.save(puffer)

        geladen = excel_to_projects(puffer.getvalue())
        assert len(geladen) == 1
        assert geladen[0].pacht_modus == PachtModus.FIX
        assert geladen[0].pacht_umsatzbeteiligung_pct == 0.055
        assert geladen[0].pacht_mindestpacht_eur_ha_jahr == 0.0


class TestDirektvermarktungsModus:
    def test_relativ_marktwert_berechnet_anteil(self, project, global_assumptions):
        """Im Relativ-Modus: DV-Kosten = Produktion x Marktwert(nominal) x
        Anteil - Jahr fuer Jahr exakt."""
        from engine import DirektvermarktungsModus, run_valuation

        ga = global_assumptions.model_copy(deep=True)
        ga.direktvermarktung_modus = DirektvermarktungsModus.RELATIV_MARKTWERT
        ga.direktvermarktung_pct_marktwert = 0.10

        df = run_valuation(project, ga).cashflow.data
        betrieb = df[df["jahr"] >= 1]
        erwartet = (
            betrieb["produktion_kwh"]
            * betrieb["marktwert_nominal_ct_kwh"]
            / 100.0
            * 0.10
        )
        assert np.allclose(betrieb["direktvermarktungskosten_eur"], erwartet)

    def test_absolut_bleibt_unveraendert(self, project, global_assumptions):
        """Der Standard-Modus ABSOLUT rechnet exakt wie bisher: fester
        EUR/kWh-Satz auf die erzeugte Menge."""
        from engine import run_valuation

        df = run_valuation(project, global_assumptions).cashflow.data
        betrieb = df[df["jahr"] >= 1]
        satz = project.direktvermarktungskosten_eur_mwh / 1000
        assert np.allclose(
            betrieb["direktvermarktungskosten_eur"],
            betrieb["produktion_kwh"] * satz,
        )

    def test_modus_aendert_irr(self, project, global_assumptions):
        """Ein spuerbarer Marktwert-Anteil (10 %) muss die Rendite gegen-
        ueber 1 EUR/MWh absolut deutlich druecken."""
        from engine import DirektvermarktungsModus, run_valuation

        irr_absolut = run_valuation(project, global_assumptions).kpis.equity_irr
        ga = global_assumptions.model_copy(deep=True)
        ga.direktvermarktung_modus = DirektvermarktungsModus.RELATIV_MARKTWERT
        ga.direktvermarktung_pct_marktwert = 0.10
        irr_relativ = run_valuation(project, ga).kpis.equity_irr
        assert irr_relativ < irr_absolut

    def test_yaml_roundtrip_mit_modus(self, tmp_path, global_assumptions):
        from engine import DirektvermarktungsModus
        from engine.io_yaml import (
            load_global_assumptions_yaml,
            save_global_assumptions_yaml,
        )

        ga = global_assumptions.model_copy(deep=True)
        ga.direktvermarktung_modus = DirektvermarktungsModus.RELATIV_MARKTWERT
        ga.direktvermarktung_pct_marktwert = 0.07
        pfad = tmp_path / "ga.yaml"
        save_global_assumptions_yaml(ga, pfad)
        geladen = load_global_assumptions_yaml(pfad)
        assert geladen.direktvermarktung_modus == DirektvermarktungsModus.RELATIV_MARKTWERT
        assert geladen.direktvermarktung_pct_marktwert == 0.07

    def test_excel_roundtrip_mit_modus(self, global_assumptions):
        from engine import DirektvermarktungsModus
        from engine.io_excel import (
            excel_to_global_assumptions,
            global_assumptions_to_excel,
        )

        ga = global_assumptions.model_copy(deep=True)
        ga.direktvermarktung_modus = DirektvermarktungsModus.RELATIV_MARKTWERT
        ga.direktvermarktung_pct_marktwert = 0.12
        geladen = excel_to_global_assumptions(global_assumptions_to_excel(ga))
        assert geladen.direktvermarktung_modus == DirektvermarktungsModus.RELATIV_MARKTWERT
        assert geladen.direktvermarktung_pct_marktwert == pytest.approx(0.12)
