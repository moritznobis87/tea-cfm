"""Tests fuer Finanzierung (Annuitaet/linear) und Steuer (Verlustvortrag)."""

from __future__ import annotations

from datetime import date

import pandas as pd
import pytest

from engine import TaxModus, TilgungsArt
from engine.financing import calculate_financing
from engine.tax import calculate_tax
from engine.timeline import build_timeline


class TestFinancing:
    def test_annuitaet_tilgt_kredit_vollstaendig(self):
        timeline = build_timeline(date(2027, 1, 1), 25)
        fin = calculate_financing(
            timeline,
            investitionsvolumen_eur=1_000_000.0,
            eigenkapitalquote_pct=0.2,
            fremdkapitalzins_pct=0.04,
            kreditlaufzeit_jahre=20,
            tilgungsart=TilgungsArt.ANNUITAET,
        )
        # 800.000 € Fremdkapital, nach 20 Jahren vollstaendig getilgt.
        assert fin["darlehensstand_bop_eur"].iloc[0] == pytest.approx(800_000.0)
        assert fin["darlehensstand_eop_eur"].iloc[19] == pytest.approx(0.0, abs=1e-6)
        assert fin["tilgung_eur"].sum() == pytest.approx(800_000.0)
        # Konstante Annuitaet waehrend der Laufzeit.
        schuldendienst = fin["schuldendienst_eur"].iloc[:20]
        assert schuldendienst.max() == pytest.approx(schuldendienst.min())

    def test_lineare_tilgung_konstante_raten(self):
        timeline = build_timeline(date(2027, 1, 1), 25)
        fin = calculate_financing(
            timeline, 1_000_000.0, 0.2, 0.04, 20, TilgungsArt.LINEAR
        )
        assert fin["tilgung_eur"].iloc[:20].tolist() == pytest.approx([40_000.0] * 20)
        # Zinsen fallen mit sinkendem Darlehensstand.
        assert fin["zinsen_eur"].iloc[0] > fin["zinsen_eur"].iloc[10]

    def test_nach_kreditlaufzeit_kein_schuldendienst(self):
        timeline = build_timeline(date(2027, 1, 1), 25)
        fin = calculate_financing(
            timeline, 1_000_000.0, 0.2, 0.04, 20, TilgungsArt.ANNUITAET
        )
        assert (fin["schuldendienst_eur"].iloc[20:] == 0.0).all()


def _tax_fuer(ebt_je_jahr: list[float], **kwargs) -> pd.DataFrame:
    """Hilfskonstrukt: Steuerrechnung fuer eine vorgegebene EBT-Reihe
    (Erloese = EBT, OPEX/Zinsen = 0)."""
    n = len(ebt_je_jahr)
    revenue = pd.DataFrame({"jahr": range(1, n + 1), "erloes_eur": ebt_je_jahr})
    nullen = [0.0] * n
    opex = pd.DataFrame({"jahr": range(1, n + 1), "opex_gesamt_eur": nullen})
    financing = pd.DataFrame({"jahr": range(1, n + 1), "zinsen_eur": nullen})
    defaults = dict(
        capex_total_eur=0.0,
        tax_modus=TaxModus.PAUSCHAL_AUF_EBT,
        steuersatz_pct=0.25,
        afa_nutzungsdauer_jahre=None,
        freibetrag_eur=0.0,
        verlustvortrag_verrechnungsgrenze_pct=0.75,
    )
    defaults.update(kwargs)
    return calculate_tax(revenue, opex, financing, **defaults)


class TestTax:
    def test_pauschalsteuer_ohne_verluste(self):
        tax = _tax_fuer([100.0, 100.0])
        assert tax["steuer_eur"].tolist() == pytest.approx([25.0, 25.0])

    def test_verlustvortrag_verrechnungsgrenze_75_prozent(self):
        """Jahr 1: -100 (Verlust). Jahr 2: +100 Gewinn, davon duerfen max.
        75 % mit dem Vortrag verrechnet werden -> 25 bleiben steuerpflichtig."""
        tax = _tax_fuer([-100.0, 100.0])
        assert tax["steuer_eur"].iloc[0] == pytest.approx(0.0)
        assert tax["verlustvortrag_bestand_eur"].iloc[0] == pytest.approx(100.0)
        assert tax["verlustvortrag_genutzt_eur"].iloc[1] == pytest.approx(75.0)
        assert tax["steuerliches_ergebnis_eur"].iloc[1] == pytest.approx(25.0)
        assert tax["steuer_eur"].iloc[1] == pytest.approx(25.0 * 0.25)
        # Restvortrag bleibt bestehen (zeitlich unbegrenzt vortragbar).
        assert tax["verlustvortrag_bestand_eur"].iloc[1] == pytest.approx(25.0)

    def test_afa_nur_innerhalb_nutzungsdauer(self):
        tax = _tax_fuer(
            [100.0, 100.0, 100.0],
            tax_modus=TaxModus.AFA_KOERPERSCHAFTSTEUER,
            capex_total_eur=100.0,
            afa_nutzungsdauer_jahre=2,
        )
        assert tax["afa_eur"].tolist() == pytest.approx([50.0, 50.0, 0.0])

    def test_freibetrag_nur_im_afa_modus(self):
        pauschal = _tax_fuer([100.0], freibetrag_eur=50.0)
        afa = _tax_fuer(
            [100.0],
            tax_modus=TaxModus.AFA_KOERPERSCHAFTSTEUER,
            afa_nutzungsdauer_jahre=20,
            freibetrag_eur=50.0,
        )
        assert pauschal["steuer_eur"].iloc[0] == pytest.approx(25.0)
        assert afa["steuer_eur"].iloc[0] == pytest.approx(50.0 * 0.25)

    def test_gewerbesteuer_effektiver_satz_aus_hebesatz(self):
        """Steuermesszahl (3,5%) x Hebesatz - z.B. Hebesatz 400% -> 14%."""
        tax = _tax_fuer(
            [100.0],
            tax_modus=TaxModus.GEWERBESTEUER_DE,
            afa_nutzungsdauer_jahre=20,
            gewerbesteuer_hebesatz_pct=400.0,
            gewerbesteuer_freibetrag_eur=0.0,
        )
        assert tax["steuer_eur"].iloc[0] == pytest.approx(100.0 * 0.035 * 4.0)

    def test_gewerbesteuer_anderer_hebesatz(self):
        """Ein anderer Hebesatz (z.B. 450%, verbreitet in manchen
        Gemeinden) muss sich linear auswirken."""
        tax = _tax_fuer(
            [1000.0],
            tax_modus=TaxModus.GEWERBESTEUER_DE,
            afa_nutzungsdauer_jahre=20,
            gewerbesteuer_hebesatz_pct=450.0,
            gewerbesteuer_freibetrag_eur=0.0,
        )
        assert tax["steuer_eur"].iloc[0] == pytest.approx(1000.0 * 0.035 * 4.5)

    def test_gewerbesteuer_freibetrag_wirkt(self):
        """Gesetzlicher Freibetrag (Standard 24.500 EUR) reduziert die
        Bemessungsgrundlage vor Anwendung des Satzes."""
        ohne_freibetrag = _tax_fuer(
            [30_000.0],
            tax_modus=TaxModus.GEWERBESTEUER_DE,
            afa_nutzungsdauer_jahre=20,
            gewerbesteuer_hebesatz_pct=400.0,
            gewerbesteuer_freibetrag_eur=0.0,
        )
        mit_freibetrag = _tax_fuer(
            [30_000.0],
            tax_modus=TaxModus.GEWERBESTEUER_DE,
            afa_nutzungsdauer_jahre=20,
            gewerbesteuer_hebesatz_pct=400.0,
            gewerbesteuer_freibetrag_eur=24_500.0,
        )
        assert ohne_freibetrag["steuer_eur"].iloc[0] == pytest.approx(
            30_000.0 * 0.035 * 4.0
        )
        # Bemessungsgrundlage nach Freibetrag: 30.000 - 24.500 = 5.500
        assert mit_freibetrag["steuer_eur"].iloc[0] == pytest.approx(
            5_500.0 * 0.035 * 4.0
        )
        assert mit_freibetrag["steuer_eur"].iloc[0] < ohne_freibetrag["steuer_eur"].iloc[0]

    def test_gewerbesteuer_afa_wird_beruecksichtigt(self):
        """Wie im oesterreichischen AFA-Modus mindert die Abschreibung
        die Bemessungsgrundlage."""
        tax = _tax_fuer(
            [100.0, 100.0, 100.0],
            tax_modus=TaxModus.GEWERBESTEUER_DE,
            capex_total_eur=100.0,
            afa_nutzungsdauer_jahre=2,
            gewerbesteuer_freibetrag_eur=0.0,
        )
        assert tax["afa_eur"].tolist() == pytest.approx([50.0, 50.0, 0.0])

    def test_gewerbesteuer_ohne_verlustvortrag(self):
        """Kernanforderung: die deutsche Gewerbesteuer kennt in diesem
        vereinfachten Modell KEINEN Verlustvortrag - ein Verlustjahr
        gefolgt von einem Gewinnjahr zahlt im Gewinnjahr die VOLLE
        Steuer auf das Gewinnjahr, unabhaengig vom Vorjahresverlust
        (Gegenbeispiel: der oesterreichische Modus wuerde hier einen
        Teil verrechnen, siehe test_verlustvortrag_verrechnungsgrenze
        weiter oben)."""
        tax = _tax_fuer(
            [-100.0, 100.0],
            tax_modus=TaxModus.GEWERBESTEUER_DE,
            afa_nutzungsdauer_jahre=20,
            gewerbesteuer_hebesatz_pct=400.0,
            gewerbesteuer_freibetrag_eur=0.0,
        )
        assert tax["verlustvortrag_genutzt_eur"].iloc[1] == pytest.approx(0.0)
        assert tax["steuer_eur"].iloc[1] == pytest.approx(100.0 * 0.035 * 4.0)
        # Der Verlust "verpufft" nicht rechnerisch (bleibt als Bestand
        # sichtbar), wirkt sich aber wegen der Verrechnungsgrenze 0
        # nirgends steuermindernd aus.
        assert tax["verlustvortrag_bestand_eur"].iloc[0] == pytest.approx(100.0)

    def test_gewerbesteuer_e2e_unterscheidet_sich_von_oesterreich(
        self, project, global_assumptions
    ):
        from engine import run_valuation

        ga_at = global_assumptions.model_copy(deep=True)
        ga_at.tax_modus = TaxModus.AFA_KOERPERSCHAFTSTEUER
        r_at = run_valuation(project, ga_at)

        ga_de = global_assumptions.model_copy(deep=True)
        ga_de.tax_modus = TaxModus.GEWERBESTEUER_DE
        ga_de.gewerbesteuer_hebesatz_pct = 400.0
        ga_de.gewerbesteuer_freibetrag_eur = 24_500.0
        r_de = run_valuation(project, ga_de)

        assert r_at.kpis.equity_irr != r_de.kpis.equity_irr

    def test_yaml_roundtrip_gewerbesteuer_felder(self, tmp_path):
        from engine.io_yaml import (
            load_global_assumptions_yaml,
            save_global_assumptions_yaml,
        )

        ga = load_global_assumptions_yaml("data/global_assumptions.yaml")
        ga.tax_modus = TaxModus.GEWERBESTEUER_DE
        ga.gewerbesteuer_hebesatz_pct = 425.0
        ga.gewerbesteuer_freibetrag_eur = 24_500.0
        pfad = tmp_path / "ga.yaml"
        save_global_assumptions_yaml(ga, str(pfad))
        ga2 = load_global_assumptions_yaml(str(pfad))
        assert ga2.tax_modus == TaxModus.GEWERBESTEUER_DE
        assert ga2.gewerbesteuer_hebesatz_pct == 425.0

    def test_excel_roundtrip_gewerbesteuer_felder(self):
        from engine.io_excel import (
            excel_to_global_assumptions,
            global_assumptions_to_excel,
        )
        from engine.io_yaml import load_global_assumptions_yaml

        ga = load_global_assumptions_yaml("data/global_assumptions.yaml")
        ga.tax_modus = TaxModus.GEWERBESTEUER_DE
        ga.gewerbesteuer_hebesatz_pct = 425.0
        ga.gewerbesteuer_freibetrag_eur = 24_500.0
        ga2 = excel_to_global_assumptions(global_assumptions_to_excel(ga))
        assert ga2.tax_modus == TaxModus.GEWERBESTEUER_DE
        assert ga2.gewerbesteuer_hebesatz_pct == 425.0
        assert ga2.gewerbesteuer_freibetrag_eur == 24_500.0

    def test_excel_ohne_gewerbesteuer_spalten_faellt_auf_defaults_zurueck(self):
        """Rueckwaertskompatibilitaet: aeltere exportierte Excel-Dateien
        ohne die neuen Zeilen duerfen beim Import nicht scheitern."""
        import io

        from openpyxl import load_workbook

        from engine.io_excel import (
            excel_to_global_assumptions,
            global_assumptions_to_excel,
        )
        from engine.io_yaml import load_global_assumptions_yaml

        ga = load_global_assumptions_yaml("data/global_assumptions.yaml")
        xl = global_assumptions_to_excel(ga)
        wb = load_workbook(io.BytesIO(xl))
        ws = wb["Einstellungen"]
        for zeile in list(ws.iter_rows()):
            if zeile[0].value in (
                "gewerbesteuer_hebesatz_pct", "gewerbesteuer_freibetrag_eur",
            ):
                ws.delete_rows(zeile[0].row)
        puffer = io.BytesIO()
        wb.save(puffer)
        ga2 = excel_to_global_assumptions(puffer.getvalue())
        assert ga2.gewerbesteuer_hebesatz_pct == 400.0
        assert ga2.gewerbesteuer_freibetrag_eur == 24_500.0
